import openpyxl as p
class XlData():
    def readData(self,filepath, sheetname,r1,c1):
        sheetobj= self.sheetObj(filepath,sheetname)
        return sheetobj[0].cell(row=r1,column=c1).value

    def sheetObj(self,filepath, sheetname):
        objwb=p.load_workbook(filepath)
        sheetObj=objwb.get_sheet_by_name(sheetname)
        totalrow=sheetObj.max_row
        totalcol=sheetObj.max_column
        return sheetObj,totalcol,totalrow


    def writeData(self,filepath,sheetname,r1,c1,data):
        sheetobj= self.sheetObj(filepath, sheetname)
        sheetobj[0].cell(row=r1,column=c1).value=data